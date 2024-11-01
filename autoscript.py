import openpyxl 

from datetime import datetime 

from openpyxl import Workbook 

from openpyxl.utils.exceptions import IllegalCharacterError 

from openpyxl.utils import get_column_letter 

# Function to check for special characters except quotes and parenthesis 

def has_special_characters_except_quotes_and_parenthesis(s): 

    for char in s: 

        if char not in ['"', '(', ')'] and (not char.isalnum() and char != ' '): 

            return True 

    return False 

# Function to save error cell locations to an Excel file 

def save_error_cells_to_excel(error_cell_locations_dict, file_path, selected_sheet): 

    workbook = Workbook() 

    sheet = workbook.active 

    sheet.title = "Error Cell Locations" 

    sheet.cell(row=1, column=1, value="Column Name") 

    sheet.cell(row=1, column=2, value="Row") 

    sheet.cell(row=1, column=3, value="Column") 

    sheet.cell(row=1, column=4, value="Cell Data") 

 

    for col_name, error_cells in error_cell_locations_dict.items(): 

        col_number = None 

        for col_idx, cell in enumerate(selected_sheet[1], start=1): 

            if cell.value == col_name: 

                col_number = col_idx 

                break 

 

        if col_number is None: 

            print(f"Column '{col_name}' not found in the sheet '{selected_sheet.title}'. Skipping...") 

            continue 

 

        for i, (row, cell_data) in enumerate(error_cells, start=2): 

            sheet.cell(row=i, column=1, value=col_name) 

            sheet.cell(row=i, column=2, value=row) 

            sheet.cell(row=i, column=3, value=col_number) 

            sheet.cell(row=i, column=4, value=cell_data if cell_data else "Blank") 

 

    try: 

        workbook.save(file_path) 

        print(f"Error cell locations saved to {file_path}") 

    except Exception as e: 

        print(f"Error saving error cell locations to {file_path}: {e}") 

 

# Function to validate date format 

def is_valid_date_format(date_string, accepted_date_formats): 

    for date_format in accepted_date_formats: 

        try: 

            datetime.strptime(date_string, date_format) 

            return True 

        except ValueError: 

            pass 

    return False 

 

# Function to apply default rules based on column name 

def apply_default_rules(column_name): 

    default_date_format = None 

    if "date" in column_name.lower(): 

        default_date_format = '%d-%m-%Y' 

    return default_date_format 

 

def check_special_characters_in_column(sheet, col_number, column_name, metadata_type, accepted_date_formats): 

    special_char_count = 0 

    error_cell_locations = [] 

 

    default_date_format = apply_default_rules(column_name) 

 

    for i in range(2, sheet.max_row + 1): 

        cell_value = sheet.cell(row=i, column=col_number).value 

        cell_value = str(cell_value).strip() if cell_value else "" 

 

        if default_date_format and is_valid_date_format(cell_value, default_date_format) and i != 2: 

            continue 

 

        if metadata_type == "Percent": 

            if "%" in cell_value: 

                cell_value = cell_value.rstrip('%') 

                if cell_value.replace('.', '', 1).isdigit(): 

                    continue 

        elif metadata_type == "ID": 

            if not cell_value.isalnum(): 

                special_char_count += 1 

                error_cell_locations.append((i, cell_value)) 

                if metadata_type == "Date": 

                    accepted_date_formats = [ 

                        '%d-%m-%Y',  # dd-mm-yyyy 

                        '%m-%Y',  # mm-yyyy 

                        '%Y',  # yyyy 

                        '%Y-%m',  # yyyy-mm 

                        '%Y/%m/%d',  # yyyy/mm/dd 

                        '%Y-%m-%d',  # yyyy-mm-dd 

                        '%d/%m/%Y',  # dd/mm/yyyy 

                    ] 

                    if not is_valid_date_format(cell_value, accepted_date_formats): 

                        special_char_count += 1 

        elif metadata_type == "Text": 

            if has_special_characters_except_quotes_and_parenthesis(cell_value): 

                if '"' in cell_value and cell_value.count('"') == 2 and '(' in cell_value and ')' in cell_value: 

                    continue 

                special_char_count += 1 

                error_cell_locations.append((i, cell_value)) 

        elif metadata_type == "Int": 

            if not cell_value.replace('.', '', 1).isdigit(): 

                special_char_count += 1 

                error_cell_locations.append((i, cell_value)) 

 

    return special_char_count, error_cell_locations 

 

def main(): 

    # Step 1: Get input file path 

    file_path = input("Enter the path to the input Excel file: ") 

 

    # Load the workbook 

    try: 

        workbook = openpyxl.load_workbook(file_path) 

    except Exception as e: 

        print("Error loading Excel file.") 

        return 

 

    # Step 2: Select a sheet for error checking 

    sheet_names = workbook.sheetnames 

    print("Sheet No - Sheet Name:") 

    for i, name in enumerate(sheet_names, start=1): 

        print(f"{i}. {name}") 

 

    try: 

        selected_sheet_index = int(input("Enter the sheet number you want to check: ")) - 1 

        selected_sheet = workbook[sheet_names[selected_sheet_index]] 

        selected_sheet_name = sheet_names[selected_sheet_index] 

    except (ValueError, IndexError): 

        print("Invalid sheet number.") 

        return 

 

    # Display number of rows and columns in the selected sheet 

    num_rows = selected_sheet.max_row 

    num_columns = selected_sheet.max_column 

    print(f"Selected sheet '{selected_sheet_name}' has {num_rows} rows and {num_columns} columns.") 

 

    # Option to select sheet for metadata inputs 

    print("Do you want to provide metadata inputs for columns from a different sheet?") 

    print("1. Yes, provide metadata inputs for a different sheet.") 

    print("2. No, provide metadata inputs for the selected sheet.") 

    choice = input("Enter your choice (1 or 2): ") 

 

    if choice == "1": 

        print("Sheet No - Sheet Name:") 

        for i, name in enumerate(sheet_names, start=1): 

            print(f"{i}. {name}") 

 

        try: 

            metadata_sheet_index = int(input("Enter the sheet number for metadata inputs: ")) - 1 

            metadata_sheet = workbook[sheet_names[metadata_sheet_index]] 

        except (ValueError, IndexError): 

            print("Invalid sheet number for metadata inputs.") 

            return 

    else: 

        metadata_sheet = selected_sheet 

 

 

    # Get metadata input from the user 

    column_metadata = {} 

    for i in range(2, metadata_sheet.max_row + 1): 

        header = metadata_sheet.cell(row=i, column=1).value 

        input_type = metadata_sheet.cell(row=i, column=2).value 

        mandatory = metadata_sheet.cell(row=i, column=3).value 

 

        if header and input_type and mandatory: 

            column_metadata[header] = {"input": input_type, "mandatory": mandatory} 

 

    # Check special characters based on user-defined metadata inputs for each column 

    special_char_counts = {} 

    blank_cell_counts = {} 

    error_cell_locations_dict = {} 

 

    for col_number, column_name in enumerate(selected_sheet[1], start=1): 

        if column_name.value in column_metadata: 

            metadata_type = column_metadata[column_name.value]["input"] 

            accepted_date_formats = [ 

                '%d-%m-%Y',    # dd-mm-yyyy 

                '%m-%Y',       # mm-yyyy 

                '%Y',          # yyyy 

                '%Y-%m',       # yyyy-mm 

                '%Y/%m/%d',    # yyyy/mm/dd 

                '%Y-%m-%d',    # yyyy-mm-dd 

                '%d/%m/%Y',    # dd/mm/yyyy 

            ] 

            special_char_count, error_cell_locations = check_special_characters_in_column( 

                selected_sheet, col_number, column_name.value, metadata_type, accepted_date_formats 

            ) 

            special_char_counts[column_name.value] = special_char_count 

            error_cell_locations_dict[column_name.value] = error_cell_locations 

 

            blank_cell_count = 0 

            for i in range(2, selected_sheet.max_row + 1): 

                cell_value = selected_sheet.cell(row=i, column=col_number).value 

                cell_value = str(cell_value).strip() if cell_value else "" 

                if cell_value == "": 

                    blank_cell_count += 1 

            blank_cell_counts[column_name.value] = blank_cell_count 

 

    # Display the result for each column 

    print(f"The number of errors in the columns of sheet '{selected_sheet_name}' are:") 

    for col_name, error_count in special_char_counts.items(): 

        print(f"  {col_name} - {error_count} no of errors") 

 

    print("Detailed error counts in each column:") 

    for col_name, blank_count in blank_cell_counts.items(): 

        print(f"  {col_name}:") 

        print(f"    Blank cells - {blank_count}") 

        print(f"    Other - {special_char_counts[col_name]}") 

 

    # Option to save error cell locations to a new file 

    save_errors = input("Do you want to save error cell locations to a file? (yes/no): ") 

    if save_errors.lower() == "yes": 

        file_path = input("Enter the path to save the error cell locations Excel file: ") 

        save_error_cells_to_excel(error_cell_locations_dict, file_path, selected_sheet) 

        print("Error cell locations saved.") 

 

 

if __name__ == "__main__": 

    main() 
  